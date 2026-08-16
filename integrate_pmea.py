"""
Refresh the Pennsylvania (PMEA) prescribed-list flags in the band database.

Source: PMEA Selective Music List — Band (public Google Sheet, linked from
https://www.pmea.net/adjudication/). The sheet is the authoritative current
list, so the `PA` column is rebuilt from it rather than merged into it.

Updates WindBand_Repertoire_Database.xlsx in place:
  PA               rebuilt from the current SML
  PMEA Grade       new column — PMEA's own grade level (1-6)
  State List Count recomputed as the sum of the 22 state columns
  Street Cred      adjusted for the new state count (+0.01 per state)

Writes two review files so nothing is dropped silently:
  pmea_review_unmatched.csv  SML entries with no row in the band database
  pmea_review_dropped.csv    rows that were flagged PA but are not on the current list

Usage:
    python integrate_pmea.py            # fetch, report, write
    python integrate_pmea.py --dry-run  # fetch and report only
"""

import argparse
import os
import re
import sys

import pandas as pd
import requests
from openpyxl import load_workbook

BAND_FILE = "WindBand_Repertoire_Database.xlsx"
SHEET = "Band Originals"

SML_SHEET_ID = "1GJ2T7P-iAyQ6_xIDITiN9D3gEx8KHOfFdZrFPIzkl2M"
SML_GID = "540024496"
SML_URL = (
    f"https://docs.google.com/spreadsheets/d/{SML_SHEET_ID}"
    f"/export?format=csv&gid={SML_GID}"
)
SML_CACHE = "pmea_band_sml.csv"

# The 22 state prescribed-list columns that feed State List Count.
STATE_COLS = [
    "AR", "FL", "GA", "IA", "ID", "IN", "KS", "KY", "LA", "MA", "MD",
    "MI", "MN", "NC", "OK", "OR", "PA", "SC", "TN", "TX", "UT", "WV",
]

STATE_LIST_WEIGHT = 0.01  # Street Cred contribution per state list


# ---------------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------------

# The two sources title the same piece several different ways, so each title
# expands into a set of variants and a match on any one of them counts:
#
#   opus numbers    SML "Divertimento for Band, Op. 42" / DB "Divertimento for Band"
#   parentheticals  SML "Satiric Dances (for a Comedy by Aristophanes)"
#                   DB  "Second Suite in F (any 2 mvts.)"
#   inversion       SML "Candide, Overture to" / DB "Overture to Candide"
#   articles        SML "Musical Toast, A" / DB "A Musical Toast"

#   genre suffix    SML "Chester Overture" / DB "Chester"
#   inner articles  SML "Suite in Minor Mode" / DB "Suite in a Minor Mode"

_ARTICLES = ("the ", "a ", "an ")
_OPUS_RE = re.compile(r",?\s*(op\.?|opus|no\.?)\s*\d+.*$", re.I)
_PAREN_RE = re.compile(r"[\(\[][^\)\]]*[\)\]]")
_INNER_ARTICLE_RE = re.compile(r"\b(the|a|an)\b", re.I)
_GENRE_SUFFIX_RE = re.compile(
    r"\s+(overture|suite|march|fantasy|fanfare)?"
    r"(\s+for\s+(band|winds|concert\s+band|symphonic\s+band))?\s*$", re.I)


def _squash(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", text.lower())


def _strip_article(text: str) -> str:
    for article in _ARTICLES:
        if text.startswith(article):
            return text[len(article):]
    return text


def title_variants(value) -> set:
    """All plausible normalized spellings of one title."""
    base = str(value).strip().lower()
    forms = {base}

    for form in (base, _PAREN_RE.sub(" ", base)):
        forms.add(form)
        forms.add(_OPUS_RE.sub("", form))

    # Trailing clause after a comma, inverted back to natural order.
    for form in list(forms):
        if "," in form:
            head, tail = form.rsplit(",", 1)
            head, tail = head.strip(), tail.strip()
            if head and tail:
                forms.add(head)
                forms.add(f"{tail} {head}")

    # A bare genre suffix is often present on one side only. Only drop it when
    # something substantial remains, so "Overture" itself survives intact.
    for form in list(forms):
        trimmed = _GENRE_SUFFIX_RE.sub("", form).strip()
        if trimmed and trimmed != form and len(_squash(trimmed)) >= 5:
            forms.add(trimmed)

    for form in list(forms):
        forms.add(_INNER_ARTICLE_RE.sub(" ", form))

    return {v for v in (_squash(_strip_article(f.strip())) for f in forms) if v}


def norm_surname(value) -> str:
    """Both sources lead with the surname; first names differ (initial vs full)."""
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return ""
    head = text.split(",")[0] if "," in text else text.rsplit(" ", 1)[-1]
    return re.sub(r"[^a-z]", "", head.lower())


def keys_of(title, composer) -> set:
    """Composer surname is the anchor; the title may take several forms."""
    surname = norm_surname(composer)
    if not surname:
        return set()
    return {f"{t}|{surname}" for t in title_variants(title)}


# ---------------------------------------------------------------------------
# Load
# ---------------------------------------------------------------------------

def load_sml(refresh: bool) -> pd.DataFrame:
    if refresh or not os.path.exists(SML_CACHE):
        print(f"Fetching PMEA band SML from Google Sheets...")
        resp = requests.get(SML_URL, timeout=60)
        resp.raise_for_status()
        with open(SML_CACHE, "wb") as fh:
            fh.write(resp.content)
    else:
        print(f"Using cached {SML_CACHE}")

    sml = pd.read_csv(SML_CACHE)
    # The title header carries the list's own row count ("TITLE  5,407").
    sml = sml.rename(columns={sml.columns[0]: "Title"})
    sml.columns = [c.strip() for c in sml.columns]
    sml = sml[sml["Title"].notna()].copy()
    sml["keys"] = [keys_of(t, c) for t, c in zip(sml["Title"], sml["COMPOSER"])]
    sml["PMEA Grade"] = pd.to_numeric(sml["PMEA GRADE LEVEL"], errors="coerce")
    return sml


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true",
                    help="report the diff without writing the workbook")
    ap.add_argument("--refresh", action="store_true",
                    help="re-download the SML even if a local copy exists")
    args = ap.parse_args()

    sml = load_sml(refresh=args.refresh)
    band = pd.read_excel(BAND_FILE, sheet_name=SHEET)
    print(f"SML entries: {len(sml)}   band rows: {len(band)}")

    band["keys"] = [keys_of(t, c) for t, c in zip(band["Title"], band["Composer"])]

    # PMEA grades a title once; a title can appear several times in the band
    # database (different arrangements), and all of those rows are on the list.
    # Where PMEA lists several gradings of one title, take the lowest.
    grade_by_key = {}
    sml_keys = set()
    for keys, grade in zip(sml["keys"], sml["PMEA Grade"]):
        sml_keys |= keys
        if pd.notna(grade):
            for k in keys:
                grade_by_key[k] = min(grade_by_key.get(k, grade), grade)

    band_keys = set().union(*band["keys"]) if len(band) else set()

    new_pa = band["keys"].map(lambda ks: bool(ks & sml_keys))
    old_pa = band["PA"].astype(bool)

    added = int((new_pa & ~old_pa).sum())
    dropped = int((~new_pa & old_pa).sum())
    matched_sml = int(sml["keys"].map(lambda ks: bool(ks & band_keys)).sum())

    print(f"\nSML entries matched to a band row : {matched_sml} / {len(sml)}")
    print(f"Band rows on the current PMEA list: {int(new_pa.sum())} "
          f"(was {int(old_pa.sum())})")
    print(f"  newly flagged : +{added}")
    print(f"  dropped       : -{dropped}")

    # Review files — the dropped rows in particular deserve a human look, since
    # they may be holdovers from an earlier edition of the list.
    unmatched = ~sml["keys"].map(lambda ks: bool(ks & band_keys))
    sml.loc[unmatched, ["Title", "COMPOSER", "PMEA Grade", "PUBLISHER"]] \
        .to_csv("pmea_review_unmatched.csv", index=False)
    band.loc[~new_pa & old_pa, ["Title", "Composer", "Arranger", "Grade"]] \
        .to_csv("pmea_review_dropped.csv", index=False)
    print("\nWrote pmea_review_unmatched.csv, pmea_review_dropped.csv")

    # Recompute the two derived columns that depend on PA.
    states = band[STATE_COLS].astype(bool).copy()
    states["PA"] = new_pa
    new_count = states.sum(axis=1)
    old_count = band["State List Count"]
    new_cred = (band["Street Cred"]
                + STATE_LIST_WEIGHT * (new_count - old_count)).round(4)

    print(f"State List Count changed on {int((new_count != old_count).sum())} rows")

    if args.dry_run:
        print("\n--dry-run: workbook not modified")
        return 0

    # Update in place so existing formatting and untouched columns survive.
    wb = load_workbook(BAND_FILE)
    ws = wb[SHEET]
    header = {ws.cell(row=1, column=i).value: i
              for i in range(1, ws.max_column + 1)}

    grade_col = header.get("PMEA Grade")
    if grade_col is None:
        grade_col = ws.max_column + 1
        ws.cell(row=1, column=grade_col, value="PMEA Grade")

    pa_col = header["PA"]
    count_col = header["State List Count"]
    cred_col = header["Street Cred"]

    def lookup_grade(keys):
        grades = [grade_by_key[k] for k in keys if k in grade_by_key]
        return min(grades) if grades else None

    pmea_grade = band["keys"].map(lookup_grade)
    for i in range(len(band)):
        row = i + 2  # header occupies row 1
        ws.cell(row=row, column=pa_col, value=bool(new_pa.iat[i]))
        ws.cell(row=row, column=count_col, value=int(new_count.iat[i]))
        ws.cell(row=row, column=cred_col, value=float(new_cred.iat[i]))
        grade = pmea_grade.iat[i]
        ws.cell(row=row, column=grade_col,
                value=None if pd.isna(grade) else int(grade))

    wb.save(BAND_FILE)
    print(f"\nUpdated {BAND_FILE} (PA, PMEA Grade, State List Count, Street Cred)")
    return 0


if __name__ == "__main__":
    sys.exit(main())

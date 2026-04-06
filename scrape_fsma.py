"""
FSMA MPA Results Scraper
========================
Downloads MPA results XLSX files for all available years from
flmusiced.org/MPAOnline/PublicReports/ExportMPAResults.aspx

Then pivots each into the flat CSV format and merges into fsma_results.csv.
"""

import requests
from bs4 import BeautifulSoup
import time
import logging
import sys
import csv
from pathlib import Path
from collections import defaultdict
import openpyxl
import tempfile

# ── Configuration ─────────────────────────────────────────────────────────────

EXPORT_URL = "https://flmusiced.org/MPAOnline/PublicReports/ExportMPAResults.aspx"

# All available school years (2010-2011 through 2026-2027)
ALL_YEARS = [
    "2010-2011", "2011-2012", "2012-2013", "2013-2014", "2014-2015",
    "2015-2016", "2016-2017", "2017-2018", "2018-2019", "2019-2020",
    "2020-2021", "2021-2022", "2022-2023", "2023-2024", "2024-2025",
    "2025-2026", "2026-2027",
]

REQUEST_DELAY = 3.0  # seconds between requests — be polite
BACKOFF_DELAY = 30
MAX_RETRIES = 3

XLSX_DIR = Path("fsma_xlsx")
OUTPUT_FILE = Path("fsma_results.csv")
CHECKPOINT_FILE = Path("fsma_checkpoint.txt")
LOG_FILE = Path("fsma_scraper.log")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko)",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Referer": EXPORT_URL,
}

# ── Logging ───────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ── Rating map ────────────────────────────────────────────────────────────────

RATING_MAP = {
    "Superior": 1,
    "Excellent": 2,
    "Good": 3,
    "Fair": 4,
    "Poor": 5,
}

FIELDNAMES = [
    "school_year", "component", "school", "fsma_school_id",
    "ensemble", "directors", "level", "mpa_type",
    "grade_level", "classification", "student_count",
    "final_rating", "final_rating_num",
    "concert_j1", "concert_j1_rating", "concert_j1_num",
    "concert_j2", "concert_j2_rating", "concert_j2_num",
    "concert_j3", "concert_j3_rating", "concert_j3_num",
    "sr_j1", "sr_j1_rating", "sr_j1_num",
    "music_j1", "music_j1_rating", "music_j1_num",
    "music_j2", "music_j2_rating", "music_j2_num",
    "other_judges",
]

# ── ASP.NET form helpers ─────────────────────────────────────────────────────

def get_viewstate(session):
    """GET the export page and extract ASP.NET hidden fields."""
    resp = session.get(EXPORT_URL, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    fields = {}
    for name in ("__VIEWSTATE", "__VIEWSTATEGENERATOR", "__EVENTVALIDATION"):
        tag = soup.find("input", {"name": name})
        if tag:
            fields[name] = tag.get("value", "")

    # Also grab available years from the dropdown for validation
    ddl = soup.find("select", {"name": "ctl00$Content$ddlYear"})
    available = []
    if ddl:
        for opt in ddl.find_all("option"):
            val = opt.get("value", "").strip()
            if val and val != "Select School Year...":
                available.append(val)

    return fields, available


def download_year(session, year, viewstate_fields):
    """POST the export form for a given year, return XLSX bytes or None."""
    payload = dict(viewstate_fields)
    payload["ctl00$Content$ddlYear"] = year + " "  # trailing space matches captured request
    payload["ctl00$Content$Button1"] = "Export"
    payload["__EVENTTARGET"] = ""
    payload["__EVENTARGUMENT"] = ""
    payload["__LASTFOCUS"] = ""

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.post(EXPORT_URL, data=payload, headers=HEADERS, timeout=60)
            if resp.status_code == 200:
                content_type = resp.headers.get("Content-Type", "")
                if "spreadsheet" in content_type or "excel" in content_type or resp.content[:4] == b"PK\x03\x04":
                    return resp.content
                else:
                    # Might be an HTML error page or empty result
                    if len(resp.content) < 500:
                        log.warning(f"  Small response ({len(resp.content)} bytes) — likely no data")
                        return None
                    # Could be a new page with updated viewstate — try re-fetching
                    log.warning(f"  Got HTML instead of XLSX (attempt {attempt})")
                    if attempt < MAX_RETRIES:
                        viewstate_fields, _ = get_viewstate(session)
                        payload.update(viewstate_fields)
                        time.sleep(BACKOFF_DELAY)
            elif resp.status_code in (429, 503):
                log.warning(f"  Rate-limited ({resp.status_code}) — sleeping {BACKOFF_DELAY}s")
                time.sleep(BACKOFF_DELAY)
            else:
                log.warning(f"  HTTP {resp.status_code}")
                return None
        except requests.RequestException as e:
            log.error(f"  Request error (attempt {attempt}): {e}")
            time.sleep(BACKOFF_DELAY)

    return None


# ── XLSX → rows pivot ────────────────────────────────────────────────────────

def pivot_xlsx(xlsx_path):
    """Read an FSMA XLSX and return pivoted row dicts."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    groups = defaultdict(lambda: {"meta": None, "judges": []})

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = [str(c).strip() if c else "" for c in row]
        if len(vals) < 15:
            continue

        key = (vals[2], vals[4], vals[6], vals[7])
        if groups[key]["meta"] is None:
            groups[key]["meta"] = {
                "school_year": vals[0],
                "component": vals[1],
                "school": vals[2],
                "fsma_school_id": vals[3],
                "ensemble": vals[4],
                "directors": vals[5],
                "level": vals[6],
                "mpa_type": vals[7],
                "grade_level": vals[8],
                "classification": vals[9],
                "student_count": vals[10],
                "final_rating": vals[14],
                "final_rating_num": RATING_MAP.get(vals[14], ""),
            }
        groups[key]["judges"].append({
            "name": vals[11],
            "category": vals[12],
            "rating": vals[13],
            "rating_num": RATING_MAP.get(vals[13], ""),
        })

    wb.close()

    rows_out = []
    for key, group in sorted(groups.items()):
        row = dict(group["meta"])
        judges = group["judges"]

        concert = [j for j in judges if j["category"] == "Concert"]
        sr = [j for j in judges if j["category"] == "Sight-Reading"]
        music = [j for j in judges if j["category"] == "Music"]
        other = [j for j in judges if j["category"] not in ("Concert", "Sight-Reading", "Music")]

        for idx, j in enumerate(concert[:3], start=1):
            row[f"concert_j{idx}"] = j["name"]
            row[f"concert_j{idx}_rating"] = j["rating"]
            row[f"concert_j{idx}_num"] = j["rating_num"]

        if sr:
            row["sr_j1"] = sr[0]["name"]
            row["sr_j1_rating"] = sr[0]["rating"]
            row["sr_j1_num"] = sr[0]["rating_num"]

        for idx, j in enumerate(music[:2], start=1):
            row[f"music_j{idx}"] = j["name"]
            row[f"music_j{idx}_rating"] = j["rating"]
            row[f"music_j{idx}_num"] = j["rating_num"]

        if other:
            row["other_judges"] = "; ".join(
                f"{j['name']}:{j['category']}:{j['rating']}" for j in other
            )

        for f in FIELDNAMES:
            row.setdefault(f, "")

        rows_out.append(row)

    return rows_out


# ── Checkpoint ───────────────────────────────────────────────────────────────

def load_checkpoint():
    done = set()
    if CHECKPOINT_FILE.exists():
        for line in CHECKPOINT_FILE.read_text().splitlines():
            done.add(line.strip())
    return done


def save_checkpoint(year):
    with open(CHECKPOINT_FILE, "a") as f:
        f.write(f"{year}\n")


# ── Main ─────────────────────────────────────────────────────────────────────

def main(years=None):
    XLSX_DIR.mkdir(exist_ok=True)
    done = load_checkpoint()
    session = requests.Session()

    log.info("Fetching export page to get ViewState and available years...")
    viewstate_fields, available_years = get_viewstate(session)
    log.info(f"Available years on site: {available_years}")

    if years is None:
        years = available_years if available_years else ALL_YEARS

    # Skip already-completed years
    remaining = [y for y in years if y.strip() not in done]
    log.info(f"Years to scrape: {len(remaining)} (skipping {len(years) - len(remaining)} already done)")

    all_rows = []

    # Load any previously-downloaded XLSX files first
    for y in years:
        xlsx_path = XLSX_DIR / f"FSMA_{y.replace('-', '_')}.xlsx"
        if xlsx_path.exists() and y.strip() in done:
            log.info(f"Loading cached {xlsx_path.name}...")
            rows = pivot_xlsx(xlsx_path)
            all_rows.extend(rows)
            log.info(f"  → {len(rows)} ensemble records")

    # Download remaining years
    for i, year in enumerate(remaining, 1):
        log.info(f"[{i}/{len(remaining)}] Downloading {year}...")

        xlsx_bytes = download_year(session, year, viewstate_fields)

        if xlsx_bytes is None:
            log.warning(f"  → No data for {year}")
            save_checkpoint(year.strip())
            time.sleep(REQUEST_DELAY)
            continue

        # Save the raw XLSX
        xlsx_path = XLSX_DIR / f"FSMA_{year.replace('-', '_').strip()}.xlsx"
        xlsx_path.write_bytes(xlsx_bytes)
        log.info(f"  → Saved {xlsx_path.name} ({len(xlsx_bytes):,} bytes)")

        # Pivot and accumulate
        rows = pivot_xlsx(xlsx_path)
        all_rows.extend(rows)
        log.info(f"  → {len(rows)} ensemble records (running total: {len(all_rows)})")

        save_checkpoint(year.strip())

        # Refresh ViewState for next request
        try:
            viewstate_fields, _ = get_viewstate(session)
        except Exception as e:
            log.warning(f"  ViewState refresh failed: {e}")

        time.sleep(REQUEST_DELAY)

    # Write merged CSV
    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(all_rows)

    log.info(f"\nDone. {len(all_rows)} total ensemble records written to {OUTPUT_FILE}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Scrape FSMA MPA results")
    parser.add_argument("--years", nargs="+", default=None,
                        help="School years to scrape (e.g. --years 2023-2024 2022-2023)")
    args = parser.parse_args()

    main(years=args.years)
